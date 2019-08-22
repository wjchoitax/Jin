import openpyxl


def main():
    # 엑셀을 읽어온다
    wb = openpyxl.load_workbook('C:\\Python\\examples\\07\\수강생_결제정보.xlsx')
    sheet = wb['결제정보']

    not_pay_list

    for row_index in range(2, sheet.max_row + 1):
        if sheet.cell(row_index, 4).value == '미결제':
            not_pay_list.append(sheet[row_index])


main()
